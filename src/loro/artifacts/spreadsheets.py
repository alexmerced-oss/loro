from csv import writer as csv_writer
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

from loro.artifacts.common import (
    ArtifactResult,
    ensure_output_dir,
    formula_safe,
    title_from_prompt,
    unique_slug,
)
from loro.artifacts.generation import SpreadsheetPayload


@dataclass(frozen=True)
class WorkbookPlan:
    title: str
    sheets: list[str]
    assumptions: list[str]


def create_spreadsheet_artifact(
    prompt: str, output_dir: Path, *, draft: SpreadsheetPayload | None = None
) -> ArtifactResult:
    title = draft.title if draft else title_from_prompt(prompt, "Loro Workbook")
    slug = unique_slug(title, "workbook")
    output_dir = ensure_output_dir(output_dir)
    xlsx_path = output_dir / f"{slug}.xlsx"
    csv_path = output_dir / f"{slug}-summary.csv"

    if draft:
        rows = [
            tuple(draft.columns),
            *(tuple(formula_safe_cell(value) for value in row) for row in draft.rows),
        ]
        data_rows: list[tuple[str, int, int]] = []
    else:
        rows = []
        data_rows = [
            ("Scope", 10, 8),
            ("Schedule", 10, 9),
            ("Risk", 10, 6),
        ]
    start_row = 4
    # Variance formulas must reference the row each value actually lands on, which
    # depends on start_row — hard-coded C2/C3/C4 pointed at the header and empty cells.
    if not draft:
        rows = [("Category", "Planned", "Actual", "Variance")]
        rows.extend(
            (category, planned, actual, f"=C{start_row + 1 + offset}-B{start_row + 1 + offset}")
            for offset, (category, planned, actual) in enumerate(data_rows)
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet["A1"] = title
    sheet["A2"] = "Prompt"
    # Prompt text is user/model controlled: a leading =, +, - or @ would otherwise be
    # stored as a live formula that executes when the workbook is opened.
    sheet["B2"] = formula_safe(prompt.strip())
    for row_index, row in enumerate(rows, start=start_row):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    if not draft:
        chart = BarChart()
        chart.title = "Planned vs Actual"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Category"
        data = Reference(sheet, min_col=2, max_col=3, min_row=start_row, max_row=start_row + 3)
        categories = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + 3)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, "F4")
    workbook.save(xlsx_path)

    # The CSV carries computed variances rather than formula text: a leading "=" in a CSV
    # cell is executed by spreadsheet apps on open.
    csv_rows: list[tuple[object, ...]] = (
        rows
        if draft
        else [
            ("Category", "Planned", "Actual", "Variance"),
            *(
                (category, planned, actual, actual - planned)
                for category, planned, actual in data_rows
            ),
        ]
    )
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        csv_writer(handle).writerows(csv_rows)

    # Validate the workbook can be opened after writing.
    load_workbook(xlsx_path, data_only=False).close()

    return ArtifactResult(
        title=title,
        kind="spreadsheet",
        paths=[xlsx_path, csv_path],
        summary=f"Created spreadsheet artifacts: {xlsx_path} and {csv_path}",
    )


def formula_safe_cell(value: str | int | float | bool | None) -> object:
    return formula_safe(value) if isinstance(value, str) else value
