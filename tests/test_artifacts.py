import json
from pathlib import Path

from openpyxl import load_workbook

from loro.artifacts.briefs import create_brief_artifact
from loro.artifacts.common import verify_provenance, write_provenance
from loro.artifacts.documents import create_document_artifact
from loro.artifacts.generation import (
    DocumentPayload,
    SpreadsheetPayload,
    document_draft,
    parse_generated_payload,
)
from loro.artifacts.presentations import create_presentation_artifact
from loro.artifacts.spreadsheets import create_spreadsheet_artifact


def test_document_artifact(tmp_path: Path) -> None:
    result = create_document_artifact("Draft onboarding guide", tmp_path)
    assert len(result.paths) == 2
    assert result.paths[0].suffix == ".md"
    assert result.paths[1].suffix == ".docx"
    assert all(path.exists() for path in result.paths)


def test_document_artifact_renders_generated_draft(tmp_path: Path) -> None:
    result = create_document_artifact(
        "ignored scaffold",
        tmp_path,
        draft=document_draft(
            DocumentPayload(
                kind="document",
                title="Useful Guide",
                body_markdown="## Answer\n\nSubstantive model-written content.",
            )
        ),
    )
    markdown = result.paths[0].read_text(encoding="utf-8")
    assert "Useful Guide" in markdown
    assert "Substantive model-written content" in markdown
    assert "deterministic MVP" not in markdown


def test_document_artifact_does_not_duplicate_model_heading(tmp_path: Path) -> None:
    result = create_document_artifact(
        "ignored scaffold",
        tmp_path,
        draft=document_draft(
            DocumentPayload(
                kind="document",
                title="Useful Guide",
                body_markdown="# Useful Guide\n\n## Answer\n\nFinished content.",
            )
        ),
    )

    markdown = result.paths[0].read_text(encoding="utf-8")
    assert markdown.count("# Useful Guide") == 1
    assert "## Answer\n\nFinished content." in markdown


def test_presentation_artifact(tmp_path: Path) -> None:
    result = create_presentation_artifact("Quarterly business review", tmp_path)
    assert any(path.suffix == ".pptx" for path in result.paths)
    assert all(path.exists() for path in result.paths)


def test_spreadsheet_artifact(tmp_path: Path) -> None:
    result = create_spreadsheet_artifact("Launch readiness tracker", tmp_path)
    workbook_path = next(path for path in result.paths if path.suffix == ".xlsx")
    workbook = load_workbook(workbook_path, data_only=False)
    assert "Summary" in workbook.sheetnames
    workbook.close()


def test_spreadsheet_artifact_renders_generated_rows_safely(tmp_path: Path) -> None:
    draft = SpreadsheetPayload(
        kind="spreadsheet",
        title="Generated Tracker",
        columns=["Item", "Value"],
        rows=[["Alpha", 3], ["=unsafe()", 4]],
    )
    result = create_spreadsheet_artifact("tracker", tmp_path, draft=draft)
    workbook_path = next(path for path in result.paths if path.suffix == ".xlsx")
    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook["Summary"]["A5"].value == "Alpha"
    assert workbook["Summary"]["A6"].value == "'=unsafe()"
    workbook.close()


def test_generated_payload_parser_validates_kind_and_shape() -> None:
    parsed = parse_generated_payload(
        'prefix {"kind":"document","title":"Guide","body_markdown":"Complete text"}',
        expected_kind="document",
    )
    assert isinstance(parsed, DocumentPayload)
    assert parsed.title == "Guide"

    try:
        parse_generated_payload(
            '{"kind":"spreadsheet","title":"Bad","columns":["A"],"rows":[[1,2]]}',
            expected_kind="spreadsheet",
        )
    except ValueError as error:
        assert "column count" in str(error)
    else:
        raise AssertionError("Expected invalid spreadsheet rows to be rejected")


def test_brief_artifact(tmp_path: Path) -> None:
    result = create_brief_artifact("Prepare for roadmap sync", tmp_path, brief_type="meeting")
    assert result.paths[0].exists()
    assert "## Next Steps" in result.paths[0].read_text(encoding="utf-8")


def test_artifact_provenance(tmp_path: Path) -> None:
    result = create_brief_artifact("Prepare for roadmap sync", tmp_path, brief_type="meeting")
    provenance = write_provenance(result=result, prompt_preview="Prepare for roadmap sync")
    assert provenance.exists()
    payload = json.loads(provenance.read_text(encoding="utf-8"))
    assert payload["generator"] == "loro.artifacts"
    assert payload["schema_version"] == "1.0"
    assert payload["artifacts"][0]["sha256"].startswith("sha256:")
    assert verify_provenance(provenance).ok


def test_artifact_provenance_detects_mutation_and_missing_artifact(tmp_path: Path) -> None:
    result = create_brief_artifact("Prepare for roadmap sync", tmp_path, brief_type="meeting")
    provenance = write_provenance(result=result, prompt_preview="Prepare for roadmap sync")

    result.paths[0].write_text("mutated", encoding="utf-8")
    mutated = verify_provenance(provenance)
    assert not mutated.ok
    assert "binding mismatch" in mutated.issues[0]

    result.paths[0].unlink()
    missing = verify_provenance(provenance)
    assert not missing.ok
    assert "artifact missing" in missing.issues[0]


def test_artifact_provenance_rejects_non_object_root(tmp_path: Path) -> None:
    provenance = tmp_path / "artifact.provenance.json"
    provenance.write_text("[]", encoding="utf-8")

    report = verify_provenance(provenance)

    assert not report.ok
    assert report.issues == ["provenance root must be an object"]
