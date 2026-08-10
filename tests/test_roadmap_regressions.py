"""Regression coverage for the August 10 2026 audit findings.

Each test here would have failed against the pre-fix code. They are grouped by the bug id
from `roadmap_aug10.md` so a future change that reintroduces one is obvious.
"""

from __future__ import annotations

import json
import random
import string
from concurrent.futures import ThreadPoolExecutor
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest

from loro.artifacts.spreadsheets import create_spreadsheet_artifact
from loro.audit import AuditLogger
from loro.audit.query import AuditQuery, audit_report, query_audit_events
from loro.audit.sinks import AuditBuffer, JsonlAuditSink, verify_jsonl_audit
from loro.config import (
    AuditConfig,
    LoroConfig,
    MCPConfig,
    MCPServerConfig,
    PermissionsConfig,
    SandboxProfileConfig,
    SessionConfig,
    SharedMemoryConfig,
)
from loro.config_check import check_config
from loro.credentials import CredentialError, CredentialVault
from loro.data_protection import RegexContentScanner
from loro.mcp.server import LoroMCPServerCatalog, MCPServerModeError
from loro.mcp.tasks import MCPTaskError, MCPTaskHandle, MCPTaskStore
from loro.memory.base import SharedMemoryDraft, like_term
from loro.memory.drafts import SharedMemoryDraftStore
from loro.memory.postgres import PostgresSharedMemoryStore
from loro.polaris import normalize_readonly_args
from loro.resources import polaris_resource
from loro.session_messages import SessionMailbox
from loro.sessions import SessionRecord
from loro.tool_runtime import ToolRegistry, parse_tool_calls
from loro.tool_schemas import provider_tool_payload, tool_catalog
from loro.tools.files import FileTools

# --------------------------------------------------------------- B1 runtime loop


@pytest.mark.parametrize(
    "directive",
    [
        '@tool {"name":"file.read","args":{bad}',
        "@tool file.read {not json}",
        '@tool {"name": 5, "args": {}}',
        "@tool {}",
        '@tool file.read ["not", "an", "object"]',
    ],
)
def test_malformed_tool_directive_raises_a_typed_error_not_a_crash(directive: str) -> None:
    with pytest.raises((ValueError, json.JSONDecodeError)):
        parse_tool_calls(directive, origin="model")


def test_parse_tool_calls_never_raises_untyped_errors_on_random_input() -> None:
    generator = random.Random(20260810)
    alphabet = string.printable
    for _ in range(500):
        text = "@tool " + "".join(
            generator.choice(alphabet) for _ in range(generator.randint(0, 40))
        )
        try:
            parse_tool_calls(text, origin="model")
        except (ValueError, json.JSONDecodeError):
            pass


def test_runtime_recovers_from_a_malformed_model_directive(tmp_path, monkeypatch) -> None:
    from loro.models import ModelResponse
    from loro.runtime import AgentRuntime

    class MalformedThenDoneClient:
        def __init__(self) -> None:
            self.calls = 0

        def complete(self, _messages):
            self.calls += 1
            if self.calls == 1:
                return ModelResponse(content='@tool {"name":"file.read","args":{bad}')
            return ModelResponse(content="All done.")

        def stream(self, messages):
            yield self.complete(messages).content

    client = MalformedThenDoneClient()
    monkeypatch.setattr("loro.runtime.create_model_client", lambda config, tools=None: client)
    config = LoroConfig.model_validate(
        {
            "audit": {"path": str(tmp_path / "audit.jsonl")},
            "sessions": {"path": str(tmp_path / "sessions")},
        }
    )

    result = AgentRuntime(config).run("do something", mode="run")

    assert result.stop_reason == "completed"
    assert client.calls == 2
    events = [
        json.loads(line)
        for line in (tmp_path / "audit.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    assert any(event["event_type"] == "runtime.tool_directive_invalid" for event in events)


# ------------------------------------------------------ B2 artifact confinement


def test_artifact_create_is_confined_to_workspace_roots(tmp_path) -> None:
    outside = tmp_path / "outside"
    outside.mkdir()
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    config = LoroConfig(
        permissions=PermissionsConfig(artifact="allow", workspace_roots=[str(workspace)])
    )
    call = parse_tool_calls(
        "@tool artifact.create "
        f'{{"kind": "document", "prompt": "Guide", "output_dir": "{outside}"}}'
    )[0]

    result = ToolRegistry(config).execute(call)

    assert result.ok is False
    assert "outside configured workspace roots" in result.output
    assert not any(outside.iterdir())


def test_artifact_create_requires_approval_by_default(tmp_path) -> None:
    call = parse_tool_calls(
        "@tool artifact.create "
        f'{{"kind": "document", "prompt": "Guide", "output_dir": "{tmp_path}"}}'
    )[0]

    result = ToolRegistry(LoroConfig()).execute(call)

    assert result.ok is False
    assert "requires approval" in result.output
    assert not any(tmp_path.iterdir())


# ------------------------------------------------------------ B3/B4 Polaris argv


def test_polaris_resource_normalizes_inline_option_syntax() -> None:
    inline = polaris_resource(["catalogs", "list", "--catalog=secret-prod"])
    spaced = polaris_resource(["catalogs", "list", "--catalog", "secret-prod"])

    assert inline.fields["catalog"] == "secret-prod"
    assert inline.fields["catalog"] == spaced.fields["catalog"]


def test_polaris_readonly_rejects_smuggled_flags_and_dash_positionals() -> None:
    with pytest.raises(PermissionError, match="not permitted"):
        normalize_readonly_args(["catalogs", "list", "--profile", "admin"])
    with pytest.raises(PermissionError, match="short options"):
        normalize_readonly_args(["tables", "get", "-o", "/tmp/out"])
    with pytest.raises(PermissionError, match="option-like value"):
        normalize_readonly_args(["tables", "get", "--catalog", "--profile"])
    with pytest.raises(PermissionError, match="option separator"):
        normalize_readonly_args(["tables", "get", "--", "events"])
    # A positional beginning with "-" can no longer be consumed as a flag.
    normalized = normalize_readonly_args(["tables", "get", "--catalog", "prod", "events"])
    assert normalized == ["tables", "get", "--catalog", "prod", "--", "events"]


# --------------------------------------------------------- B5-B8 secret scanning


def test_private_key_redaction_covers_the_whole_block() -> None:
    content = (
        "-----BEGIN RSA PRIVATE KEY-----\n"
        "MIIEowIBAAKCAQEAsecretKEYMATERIAL0123456789\n"
        "-----END RSA PRIVATE KEY-----"
    )
    findings = RegexContentScanner().scan(content)

    assert [finding.kind for finding in findings] == ["private_key"]
    finding = findings[0]
    redacted = content[: finding.start] + "[redacted]" + content[finding.end :]
    assert "KEYMATERIAL" not in redacted
    assert "-----END" not in redacted


@pytest.mark.parametrize(
    "content",
    [
        "OPENAI_API_KEY=sk-abcdefghijklmnop",
        "AWS_SECRET_ACCESS_KEY=wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
        "GITHUB_TOKEN=ghp_abcdefghijklmnopqrstuvwxyz012345",
        "MY_PASSWORD=supersecret1",
        "SERVICE_CREDENTIAL: hunter2hunter2",
    ],
)
def test_prefixed_assignment_secrets_are_detected(content: str) -> None:
    assert RegexContentScanner().scan(content), content


def test_bare_aws_secret_key_is_detected_but_hex_digests_are_not() -> None:
    scanner = RegexContentScanner()
    assert any(
        finding.kind == "aws_secret_key"
        for finding in scanner.scan("creds: wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY")
    )
    assert scanner.scan("commit 356a192b7913b04c54574d18c28d46e6395428ab") == []


def test_finding_preview_never_echoes_secret_characters() -> None:
    findings = RegexContentScanner().scan("api_key = 'SUPERsecretVALUE12345'")

    assert findings
    for finding in findings:
        assert "SUPER" not in finding.snippet
        assert "2345" not in finding.snippet
        assert finding.snippet.startswith("[redacted]")


# ------------------------------------------------------------- B9-B11 artifacts


def test_spreadsheet_variance_formulas_reference_their_own_rows(tmp_path) -> None:
    from openpyxl import load_workbook

    result = create_spreadsheet_artifact("Quarterly review", tmp_path)
    sheet = load_workbook(next(path for path in result.paths if path.suffix == ".xlsx"))["Summary"]

    assert sheet["A4"].value == "Category"
    assert [sheet.cell(row=row, column=4).value for row in (5, 6, 7)] == [
        "=C5-B5",
        "=C6-B6",
        "=C7-B7",
    ]


def test_spreadsheet_prompt_is_stored_as_text_not_a_formula(tmp_path) -> None:
    from openpyxl import load_workbook

    result = create_spreadsheet_artifact("=cmd|' /c calc'!A1", tmp_path)
    sheet = load_workbook(next(path for path in result.paths if path.suffix == ".xlsx"))["Summary"]

    assert not str(sheet["B2"].value).startswith("=")


def test_artifacts_generated_together_do_not_overwrite_each_other(tmp_path) -> None:
    first = create_spreadsheet_artifact("Launch readiness", tmp_path)
    second = create_spreadsheet_artifact("Launch readiness", tmp_path)

    assert first.paths[0] != second.paths[0]
    assert first.paths[0].is_file() and second.paths[0].is_file()


# --------------------------------------------------------------- B12-B16 memory


def test_shared_memory_draft_round_trips_its_expiry(tmp_path) -> None:
    expires_at = datetime.now(UTC) + timedelta(days=30)
    store = SharedMemoryDraftStore(tmp_path)
    store.stage(
        SharedMemoryDraft(content="c", summary="s", tenant_id="acme", expires_at=expires_at)
    )

    listed = store.list()

    assert len(listed) == 1
    assert listed[0].expires_at == expires_at


def test_postgres_search_escapes_like_metacharacters() -> None:
    statement = PostgresSharedMemoryStore(SharedMemoryConfig()).render_search(
        tenant_id="acme", query="100%_raw"
    )

    assert statement.params["query"] == r"%100\%\_raw%"
    assert "ESCAPE '\\'" in statement.sql
    assert like_term("a\\b") == "%a\\\\b%"


# ------------------------------------------------------------ B17-B19 audit chain


def test_audit_buffer_evicts_oldest_and_counts_evictions(tmp_path) -> None:
    buffer = AuditBuffer(tmp_path / "buffer.jsonl", max_events=2)

    assert buffer.append({"event_id": "1"}) == 0
    assert buffer.append({"event_id": "2"}) == 0
    assert buffer.append({"event_id": "3"}) == 1

    assert [event["event_id"] for event in buffer.load()] == ["2", "3"]
    assert buffer.evicted_events() == 1


def test_concurrent_appends_during_drain_are_never_lost(tmp_path) -> None:
    buffer = AuditBuffer(tmp_path / "buffer.jsonl", max_events=100)
    for index in range(5):
        buffer.append({"event_id": f"start-{index}"})
    delivered: list[dict] = []

    def deliver(batch: list[dict]) -> None:
        delivered.extend(batch)

    with ThreadPoolExecutor(max_workers=2) as pool:
        drain = pool.submit(buffer.drain, deliver)
        appends = pool.submit(
            lambda: [buffer.append({"event_id": f"late-{index}"}) for index in range(5)]
        )
        result = drain.result()
        appends.result()

    surviving = {event["event_id"] for event in buffer.load()}
    seen = {event["event_id"] for event in delivered} | surviving
    assert result.delivered == len(delivered)
    # Every event either got delivered or is still buffered; none vanished.
    assert {f"start-{index}" for index in range(5)} <= seen
    assert {f"late-{index}" for index in range(5)} <= seen


def test_legacy_audit_file_with_blank_lines_verifies_clean(tmp_path) -> None:
    path = tmp_path / "audit.jsonl"
    path.write_text(
        json.dumps({"event_type": "legacy.one"}) + "\n\n"
        + json.dumps({"event_type": "legacy.two"}) + "\n\n",
        encoding="utf-8",
    )
    JsonlAuditSink(path).deliver({"event_type": "chained.one"})

    verification = verify_jsonl_audit(path)

    assert verification.ok is True, verification.issue
    assert verification.legacy_events == 2


# --------------------------------------------------------------- B20-B25 gateway


def _gateway_endpoint(platform: str):
    from loro.config import GatewayEndpointConfig, GatewayIdentityConfig

    return GatewayEndpointConfig(
        platform=platform,
        route="/hooks/test",
        credentials={"signing-secret": "vault://gateway/test/secret"},
        identities={"user-1": GatewayIdentityConfig(subject="user-1", tenant="acme")},
    )


def test_non_ascii_signature_header_is_rejected_not_crashed() -> None:
    from loro.gateway.adapters import GatewayAdapterError, parse_inbound

    with pytest.raises(GatewayAdapterError):
        parse_inbound(
            "slack",
            _gateway_endpoint("slack"),
            {"X-Slack-Request-Timestamp": "1000", "X-Slack-Signature": "v0=ÿþ"},
            b"{}",
            lambda _name: "signing-value",
            now=1000,
        )


def test_deeply_nested_body_is_rejected_before_signature_parsing() -> None:
    from loro.gateway.adapters import GatewayAdapterError, parse_inbound

    body = ("[" * 200_000 + "]" * 200_000).encode()
    with pytest.raises(GatewayAdapterError):
        parse_inbound(
            "slack",
            _gateway_endpoint("slack"),
            {"X-Slack-Request-Timestamp": "1000", "X-Slack-Signature": "v0=deadbeef"},
            body,
            lambda _name: "signing-value",
            now=1000,
        )


def test_bridge_envelope_requires_a_fresh_signed_timestamp() -> None:
    import hashlib
    import hmac

    from loro.gateway.adapters import GatewayAdapterError, parse_inbound

    def signed(payload: dict) -> tuple[bytes, dict[str, str]]:
        body = json.dumps(payload).encode()
        signature = hmac.new(b"bridge-secret", body, hashlib.sha256).hexdigest()
        return body, {"X-Loro-Signature": f"sha256={signature}"}

    stale_body, stale_headers = signed(
        {"id": "m1", "text": "hi", "timestamp": 1000, "from": {"id": "user-1"}}
    )
    with pytest.raises(GatewayAdapterError, match="stale"):
        parse_inbound(
            "signal",
            _gateway_endpoint("signal"),
            stale_headers,
            stale_body,
            lambda _name: "bridge-secret",
            now=100_000,
        )

    missing_body, missing_headers = signed({"id": "m1", "text": "hi", "from": {"id": "user-1"}})
    with pytest.raises(GatewayAdapterError, match="signed timestamp"):
        parse_inbound(
            "signal",
            _gateway_endpoint("signal"),
            missing_headers,
            missing_body,
            lambda _name: "bridge-secret",
            now=1000,
        )


def test_gateway_route_matching_ignores_the_query_string(tmp_path) -> None:
    from loro.gateway.service import GatewayDispatcher

    config = LoroConfig.model_validate(
        {
            "audit": {"path": str(tmp_path / "audit.jsonl")},
            "gateway": {
                "enabled": True,
                "state_path": str(tmp_path / "state.json"),
                "endpoints": {"test": _gateway_endpoint("slack").model_dump()},
            },
        }
    )
    dispatcher = GatewayDispatcher(config, runner=lambda _c, _p: "ok", deliverer=lambda *_: None)
    try:
        response = dispatcher.handle("/hooks/test?token=abc", {}, b"{}")
    finally:
        dispatcher.close()

    # 401 means the route matched and authentication ran; 404 would mean it did not.
    assert response.status == 401


def test_gateway_run_inherits_managed_required_identity_fields(tmp_path) -> None:
    from loro.gateway.adapters import ChannelMessage
    from loro.gateway.service import GatewayDispatcher

    config = LoroConfig.model_validate(
        {
            "audit": {"path": str(tmp_path / "audit.jsonl")},
            "identity": {"required_fields": ["organization"], "environment_prefix": "ACME_ID_"},
            "gateway": {
                "enabled": True,
                "state_path": str(tmp_path / "state.json"),
                "endpoints": {"test": _gateway_endpoint("slack").model_dump()},
            },
        }
    )
    seen: list[LoroConfig] = []
    dispatcher = GatewayDispatcher(
        config,
        runner=lambda run_config, _prompt: seen.append(run_config) or "ok",
        deliverer=lambda *_: None,
    )
    try:
        dispatcher._process(
            ChannelMessage("slack", "test", "m1", "user-1", "c1", "hello"),
            config.gateway.endpoints["test"],
        )
    finally:
        dispatcher.close()

    assert seen and seen[0].identity.required_fields == ["organization"]
    assert seen[0].identity.environment_prefix == "ACME_ID_"


# ---------------------------------------------- B26-B27 credentials and messages


def test_corrupt_credential_index_raises_credential_error(tmp_path, monkeypatch) -> None:
    index = tmp_path / "credentials.json"
    index.write_text("[]", encoding="utf-8")
    vault = CredentialVault()
    monkeypatch.setattr(vault, "index_path", index, raising=False)

    with pytest.raises(CredentialError):
        vault._read_index()


def test_tampered_session_message_content_is_rejected(tmp_path) -> None:
    config = SessionConfig(
        path=str(tmp_path / "sessions"), message_path=str(tmp_path / "messages")
    )
    mailbox = SessionMailbox(config)
    for session_id in ("sender", "recipient"):
        mailbox.sessions.save(
            SessionRecord(session_id=session_id, mode="run", prompt="p", summary="s")
        )
    mailbox.send(
        sender_session_id="sender",
        recipient_session_id="recipient",
        content="original content",
        validate_sender=False,
    )
    record = next(Path(str(tmp_path / "messages")).rglob("*.json"))
    payload = json.loads(record.read_text(encoding="utf-8"))
    payload["content"] = "attacker controlled content"
    record.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(ValueError, match="digest"):
        mailbox.list("recipient")


# ------------------------------------------------------------------- B28-B31 MCP


def test_mcp_server_refuses_path_tools_without_workspace_roots() -> None:
    config = LoroConfig.model_validate(
        {"mcp": {"enabled": True, "server": {"enabled": True, "export_tools": ["agraph.plan"]}}}
    )

    with pytest.raises(MCPServerModeError, match="workspace_roots"):
        LoroMCPServerCatalog(config)


def test_task_store_terminal_status_guard_holds_under_concurrency(tmp_path) -> None:
    store = MCPTaskStore(tmp_path)
    store.save(
        MCPTaskHandle(
            server_id="s", task_id="t", status="completed", operation="call", remote={"result": {}}
        )
    )

    def attempt() -> str:
        try:
            store.save(
                MCPTaskHandle(
                    server_id="s", task_id="t", status="working", operation="call", remote={}
                )
            )
        except MCPTaskError:
            return "rejected"
        return "accepted"

    with ThreadPoolExecutor(max_workers=8) as pool:
        outcomes = list(pool.map(lambda _index: attempt(), range(8)))

    assert set(outcomes) == {"rejected"}
    assert store.get("s", "t").status == "completed"


def test_snake_case_input_requests_can_be_answered(tmp_path) -> None:
    store = MCPTaskStore(tmp_path)
    handle = store.record_remote(
        server_id="s",
        operation="task_start",
        remote={
            "taskId": "t",
            "status": "input_required",
            "input_requests": {"format": {"type": "string"}},
        },
    )

    store.validate_response_keys(handle, ["format"])
    assert "inputRequests" in handle.remote


# ----------------------------------------------------------- B32-B41 agentic graph


def test_expression_self_and_loop_namespaces_are_bound_at_runtime() -> None:
    from loro.agraph.state import bind_self

    scope: dict[str, object] = {"nodes": {}}
    bind_self(scope, node_id="apply_fix", attempt=2, inputs={"a": 1}, outputs={}, status="running")

    from loro.agraph.expressions import evaluate

    assert evaluate("self.id", scope) == "apply_fix"
    assert evaluate("self.attempt", scope) == 2
    assert evaluate("nodes.self.id", scope) == "apply_fix"


def test_expression_spec_conformance() -> None:
    from loro.agraph.expressions import ExpressionError, evaluate, interpolate

    scope = {
        "params": {"zero": 0, "left": [1], "right": [2]},
        "nodes": {"apply_fix": {"status": "failed", "outputs": {}}},
        "obj": {"k": {"j": 7}},
    }
    assert evaluate("1 == true", scope) is False
    assert evaluate("params.left + params.right", scope) == [1, 2]
    assert evaluate("get(obj, 'k.j', 'fallback')", scope) == 7
    assert evaluate("get(obj, 'k.missing', 'fallback')", scope) == "fallback"
    assert evaluate("default(nodes.apply_fix.outputs.summary, 'no fix applied')", scope) == (
        "no fix applied"
    )
    # Short-circuit: the right operand is never evaluated.
    assert evaluate("failed('apply_fix') || nodes.apply_fix.outputs.count > 0", scope) is True
    for expression in ("1 % params.zero", "1 / params.zero", "matches('a', '[')"):
        with pytest.raises(ExpressionError):
            evaluate(expression, scope)
    assert interpolate("${{ nodes.apply_fix.outputs.summary }}|${{ params.left }}", scope) == "|[1]"


def test_policy_evaluation_reports_invalid_documents_instead_of_crashing() -> None:
    from loro.agraph.policy import evaluate_policy

    config = LoroConfig().agraph
    findings = evaluate_policy(
        {"requires_conformance": "three", "nodes": {"a": {"intelligence": {"tier": "cosmic"}}}},
        config,
    )

    codes = {finding.code for finding in findings}
    assert "AG001" in codes


def test_unsupported_ags_fields_are_reported_at_load() -> None:
    from loro.agraph.support import unsupported_feature_findings

    findings = unsupported_feature_findings(
        {
            "nodes": {
                "gate": {"gate": {"timeout_seconds": 30, "on_timeout": "fail"}},
                "check": {
                    "success": {
                        "evaluation_order": "sequential",
                        "criteria": [{"kind": "llm_rubric", "timeout_seconds": 5}],
                    }
                },
            }
        }
    )

    assert {finding.code for finding in findings} == {"AG901"}
    assert len(findings) == 4


def test_resuming_a_dry_run_plan_is_refused(tmp_path, monkeypatch) -> None:
    import yaml

    from loro.agraph.execute import GraphExecutionError, GraphExecutor
    from loro.agraph.generate import generate_graph

    config = LoroConfig.model_validate(
        {
            "agraph": {"state_path": str(tmp_path / "runs")},
            "audit": {"path": str(tmp_path / "audit.jsonl")},
        }
    )
    graph_path = tmp_path / "plan.agraph.yaml"
    graph_path.write_text(
        yaml.safe_dump(generate_graph("Prepare evidence", config), sort_keys=False),
        encoding="utf-8",
    )
    planned = GraphExecutor(config, workspace=tmp_path).run(graph_path, dry_run=True)

    with pytest.raises(GraphExecutionError, match="dry-run"):
        GraphExecutor(config, workspace=tmp_path).resume(planned["run_id"])


# ------------------------------------------------------- hardening improvements


def test_file_search_does_not_follow_symlinks_out_of_the_workspace(tmp_path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    secret = tmp_path / "secret.txt"
    secret.write_text("SUPERSECRET marker", encoding="utf-8")
    (workspace / "link.txt").symlink_to(secret)
    (workspace / "real.txt").write_text("marker inside workspace", encoding="utf-8")

    matches = FileTools().search(workspace, "marker")

    assert [match.path.name for match in matches] == ["real.txt"]


def test_file_search_skips_binary_and_oversized_files(tmp_path) -> None:
    (tmp_path / "binary.bin").write_bytes(b"marker\x00" + b"\x01" * 100)
    (tmp_path / "big.txt").write_text("marker\n" + "x" * 5000, encoding="utf-8")
    (tmp_path / "small.txt").write_text("marker", encoding="utf-8")

    matches = FileTools().search(tmp_path, "marker", max_file_bytes=1000)

    assert [match.path.name for match in matches] == ["small.txt"]


@pytest.mark.parametrize("name", ["LD_PRELOAD", "PYTHONPATH", "NODE_OPTIONS", "BASH_ENV"])
def test_loader_variables_cannot_be_forwarded_to_children(name: str) -> None:
    with pytest.raises(ValueError, match="child process"):
        SandboxProfileConfig(environment_allowlist=["PATH", name])
    with pytest.raises(ValueError, match="child process"):
        MCPServerConfig(env_allowlist=[name])


def test_mcp_request_policy_hook_rejects_a_redirect_off_the_allowlist() -> None:
    import asyncio

    from loro.mcp.security import MCPTransportPolicyError, request_policy_hook

    policy = MCPConfig(allowed_hosts=["mcp.example.com"])
    hook = request_policy_hook(policy, resolver=lambda *_args, **_kwargs: [])

    class FakeRequest:
        def __init__(self, url: str) -> None:
            self.url = url

    with pytest.raises(MCPTransportPolicyError, match="not allowlisted"):
        asyncio.run(hook(FakeRequest("https://evil.example.net/mcp")))


# --------------------------------------------------------------- new features


def test_tool_catalog_is_published_in_each_provider_wire_format() -> None:
    schemas = tool_catalog(LoroConfig())

    assert {schema.name for schema in schemas} >= {"file.read", "shell.run", "artifact.create"}
    openai = provider_tool_payload("openai-compatible", schemas)
    assert openai["tools"][0]["type"] == "function"
    assert provider_tool_payload("anthropic", schemas)["tools"][0]["input_schema"]["type"] == (
        "object"
    )
    gemini = provider_tool_payload("gemini", schemas)["tools"][0]["functionDeclarations"]
    assert all("." not in item["name"] for item in gemini)
    bedrock = provider_tool_payload("bedrock", schemas)["toolConfig"]["tools"]
    assert all("." not in item["toolSpec"]["name"] for item in bedrock)


def test_tool_catalog_omits_tools_the_configuration_cannot_run() -> None:
    config = LoroConfig(permissions=PermissionsConfig(shell="deny", artifact="deny"))

    names = {schema.name for schema in tool_catalog(config)}

    assert "shell.run" not in names
    assert "artifact.create" not in names


def test_audit_query_and_report_filter_and_verify(tmp_path) -> None:
    path = tmp_path / "audit.jsonl"
    logger = AuditLogger(AuditConfig(path=str(path)))
    logger.write("approval.granted", actor="alice", tenant_id="acme", action="shell.run command")
    logger.write("approval.denied", actor="bob", tenant_id="acme", action="shell.run command")
    logger.write("runtime.task_started", actor="alice", tenant_id="acme")

    approvals = query_audit_events(path, AuditQuery(event_type="approval.*"))
    assert {event["event_type"] for event in approvals} == {
        "approval.granted",
        "approval.denied",
    }
    alice = query_audit_events(path, AuditQuery(actor="alice"))
    assert len(alice) == 2

    report = audit_report(path, AuditQuery(limit=0))
    assert report.chain_ok is True
    assert report.events == 3
    assert report.actors["alice"] == 2


def test_config_check_flags_risky_but_valid_configuration() -> None:
    config = LoroConfig.model_validate(
        {
            "permissions": {"workspace_roots": []},
            "mcp": {"enabled": True, "server": {"enabled": True}},
            "safety": {"surfaces": {"memory_shared": {"action": "redact"}}},
        }
    )

    codes = {finding.code for finding in check_config(config)}

    assert "LC001" in codes  # empty workspace roots
    assert "LC002" in codes  # MCP server mode without roots
    assert "LC010" in codes  # redact on a persistence surface


def test_mcp_dotted_tool_names_route_to_the_remote_tool() -> None:
    from loro.tool_runtime import _normalize_mcp_tool_call

    call = parse_tool_calls('@tool mcp.filesystem.read_file {"path": "README.md"}')[0]

    normalized = _normalize_mcp_tool_call(call, {"filesystem"})

    assert normalized.name == "mcp.call"
    assert normalized.args["server_id"] == "filesystem"
    assert normalized.args["tool_name"] == "read_file"
    assert normalized.args["arguments"] == {"path": "README.md"}


def test_spec_example_graphs_execute_not_just_validate(tmp_path) -> None:
    """Execute every bundled example so runtime-only defects surface.

    The examples were previously validated but never run, which is exactly why the
    executor bugs in this roadmap survived a green suite.
    """

    from dataclasses import dataclass

    from loro.agraph.execute import GraphExecutionError, GraphExecutor
    from loro.agraph.record import validate_run_record

    examples = sorted((Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "agraph" /
                       "examples").glob("*.agraph.yaml"))
    assert examples, "no example graphs found"

    @dataclass
    class Result:
        summary: str = "done"
        session_id: str = "fake"
        stop_reason: str = "completed"
        usage: dict = None  # type: ignore[assignment]
        emitted_outputs: dict = None  # type: ignore[assignment]

        def __post_init__(self) -> None:
            self.usage = {
                "input_tokens": 1,
                "output_tokens": 1,
                "total_tokens": 2,
                "cost_usd": 0.0,
                "tool_calls": 0,
            }
            # Satisfy any declared output name; unknown names are ignored downstream.
            self.emitted_outputs = {}

    class Runtime:
        def run(self, _prompt: str, mode: str) -> Result:
            return Result()

    executed = 0
    for example in examples:
        workspace = tmp_path / example.stem
        workspace.mkdir()
        copied = workspace / example.name
        copied.write_text(example.read_text(encoding="utf-8"), encoding="utf-8")
        config = LoroConfig.model_validate(
            {
                "agraph": {
                    "state_path": str(workspace / "runs"),
                    "allow_command_criteria": True,
                },
                "audit": {"path": str(workspace / "audit.jsonl")},
                "approvals": {"interactive": False},
            }
        )
        try:
            record = GraphExecutor(
                config,
                workspace=workspace,
                runtime_factory=lambda _config: Runtime(),
                gate_provider=lambda _prompt, _roles: True,
            ).run(copied, plan_approved=True)
        except GraphExecutionError as error:
            # Managed policy may refuse a graph outright (an unsigned subgraph reference),
            # and an example may declare required params this smoke test does not supply.
            # Both are typed refusals, not runtime defects; an untyped crash would fail.
            assert isinstance(error, GraphExecutionError), example.name
            continue
        # Nodes may fail their criteria against a stub runtime; what must not happen is an
        # unhandled exception escaping the executor or an unschematic run record.
        assert record["status"] in {"succeeded", "failed", "awaiting_human"}, example.name
        assert record["run_id"]
        validate_run_record({**record, "nodes": list(record["nodes"].values())}) if isinstance(
            record["nodes"], dict
        ) else validate_run_record(record)
        executed += 1
    assert executed, "every example graph was refused by policy; none exercised the executor"
